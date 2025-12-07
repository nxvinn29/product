"""
Worker to convert PDF to Excel (.xlsx)
Uses pdfplumber to extract tables and openpyxl to create Excel file
"""
from .celery_app import celery_app
import os
import pdfplumber
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


@celery_app.task(name="pdf_to_xlsx", bind=True)
def pdf_to_xlsx(self, job_id: str, input_path: str, params: dict = None) -> dict:
    """
    Convert PDF to Excel spreadsheet
    Extracts tables from PDF and creates Excel sheets
    
    Args:
        job_id: Unique job identifier
        input_path: Path to input PDF
        params: dict with optional keys:
            - extract_text: Also extract text as separate sheets (default False)
    
    Returns:
        dict with file_path to output .xlsx file
    """
    if params is None:
        params = {}
    
    output_dir = os.path.dirname(input_path)
    output_path = os.path.join(output_dir, "output.xlsx")
    
    try:
        extract_text = params.get("extract_text", False)
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        sheet_num = 1
        
        with pdfplumber.open(input_path) as pdf:
            for page_idx, page in enumerate(pdf.pages, 1):
                # Extract tables
                tables = page.extract_tables()
                
                if tables:
                    for table_idx, table in enumerate(tables, 1):
                        sheet_name = f"Page{page_idx}_Table{table_idx}"[:31]  # Max 31 char sheet name
                        ws = wb.create_sheet(title=sheet_name)
                        
                        for row_idx, row in enumerate(table, 1):
                            for col_idx, cell in enumerate(row, 1):
                                ws.cell(row=row_idx, column=col_idx, value=cell)
                
                # Optionally extract text
                if extract_text:
                    text = page.extract_text()
                    if text:
                        sheet_name = f"Text_Page{page_idx}"[:31]
                        ws = wb.create_sheet(title=sheet_name)
                        ws.cell(row=1, column=1, value=text)
        
        wb.save(output_path)
        return {"file_path": output_path}
    
    except Exception as e:
        raise Exception(f"PDF to XLSX conversion failed: {str(e)}")
